"""Excel helpers: inspect workbooks, ensure BigW columns, preview rows."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_BIGW = [
    ("BigW Name", None),
    ("BigW URL", None),
    ("BigW Match", None),
]

# Product URL column aliases (During Days sheet often uses "URL")
URL_HEADER_ALIASES = (
    "url",
    "product url",
    "product_url",
    "dd url",
    "during days url",
    "duringdays url",
    "link",
    "product link",
)


def _header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        mapping[str(cell.value).strip().lower()] = cell.column  # 1-based
    return mapping


def _find_header_index(headers: list, *names: str) -> int | None:
    lookup = {
        str(h).strip().lower(): i
        for i, h in enumerate(headers)
        if h is not None and str(h).strip()
    }
    for n in names:
        if n.lower() in lookup:
            return lookup[n.lower()]
    return None


def _row_has_match_input(row, title_idx: int | None, url_idx: int | None) -> bool:
    if title_idx is not None:
        title = row[title_idx] if title_idx < len(row) else None
        if title is not None and str(title).strip():
            return True
    if url_idx is not None:
        url = row[url_idx] if url_idx < len(row) else None
        if url is not None and str(url).strip().lower().startswith("http"):
            return True
    if title_idx is None and url_idx is None:
        return any(c is not None and str(c).strip() for c in row)
    return False


def inspect_workbook(path: Path, sheet_name: str | None = None) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    preferred = sheet_name if sheet_name in sheets else (
        "Products" if "Products" in sheets else sheets[0]
    )
    ws = wb[preferred]
    headers_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in headers_row]

    title_idx = _find_header_index(headers, "title")
    url_idx = _find_header_index(headers, *URL_HEADER_ALIASES)

    data_rows = 0
    sample: list[dict] = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not _row_has_match_input(row, title_idx, url_idx):
            continue
        data_rows += 1
        if len(sample) < 8:
            sample.append(
                {
                    "excel_row": excel_row,
                    "cells": [
                        "" if c is None else (str(c)[:120]) for c in row[:12]
                    ],
                }
            )

    header_l = {h.lower() for h in headers if h}
    missing = []
    for name, _ in REQUIRED_BIGW:
        if name.lower() not in header_l:
            missing.append(name)
    if title_idx is None:
        missing.append("Title")

    wb.close()
    return {
        "sheets": sheets,
        "sheet": preferred,
        "headers": headers,
        "data_rows": data_rows,
        "min_excel_row": 2,
        "max_excel_row": 1 + data_rows if data_rows else 1,
        "missing_bigw_columns": missing,
        "has_title_column": title_idx is not None,
        "has_url_column": url_idx is not None,
        "sample": sample,
    }


def row_bounds(path: Path, sheet: str) -> dict[str, int]:
    """Return precise min/max Excel row numbers that have a Title and/or product URL."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    title_idx = _find_header_index(headers, "title")
    url_idx = _find_header_index(headers, *URL_HEADER_ALIASES)

    first = last = None
    count = 0
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not _row_has_match_input(row, title_idx, url_idx):
            continue
        if first is None:
            first = excel_row
        last = excel_row
        count += 1
    wb.close()
    if first is None:
        return {"min_row": 2, "max_row": 2, "data_rows": 0}
    return {"min_row": first, "max_row": last, "data_rows": count}


def _place_header(ws, headers: dict[str, int], name: str) -> int:
    """Add a header column if missing; return 1-based column index."""
    if name.lower() in headers:
        return headers[name.lower()]
    for cell in ws[1]:
        if cell.value is None:
            cell.value = name
            headers[name.lower()] = cell.column
            return cell.column
    col = ws.max_column + 1
    ws.cell(row=1, column=col, value=name)
    headers[name.lower()] = col
    return col


def ensure_bigw_columns(path: Path, sheet: str) -> list[str]:
    """
    Ensure BigW Name / BigW URL / BigW Match columns exist.
    Returns list of columns that were added.
    """
    wb = load_workbook(path)
    ws = wb[sheet]
    headers = _header_map(ws)
    added: list[str] = []

    for name, _ in REQUIRED_BIGW:
        if name.lower() in headers:
            continue
        _place_header(ws, headers, name)
        added.append(name)

    wb.save(path)
    wb.close()
    return added


def ensure_title_column(path: Path, sheet: str) -> bool:
    """Ensure a Title column exists. Returns True if it was newly added."""
    wb = load_workbook(path)
    ws = wb[sheet]
    headers = _header_map(ws)
    if "title" in headers:
        wb.close()
        return False
    _place_header(ws, headers, "Title")
    wb.save(path)
    wb.close()
    return True


def preview_rows(
    path: Path,
    sheet: str,
    start_row: int,
    end_row: int,
    max_rows: int = 50,
) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    headers = [
        str(c.value).strip() if c.value is not None else f"Col{i+1}"
        for i, c in enumerate(next(ws.iter_rows(min_row=1, max_row=1)))
    ]

    focus_names = {
        "id",
        "sku",
        "title",
        "url",
        "product url",
        "bigw name",
        "bigw url",
        "bigw match",
        "kogan name",
        "kogan url",
        "kogan match",
        "validation",
    }
    focus_idx = [
        i for i, h in enumerate(headers) if h.lower() in focus_names or i < 2
    ]
    seen = set()
    focus_idx = [i for i in focus_idx if not (i in seen or seen.add(i))]

    rows_out = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if excel_row < start_row:
            continue
        if excel_row > end_row:
            break
        if len(rows_out) >= max_rows:
            break
        cells = []
        for i in focus_idx:
            val = row[i] if i < len(row) else None
            cells.append("" if val is None else str(val)[:200])
        rows_out.append({"excel_row": excel_row, "cells": cells})

    wb.close()
    return {
        "headers": [headers[i] for i in focus_idx],
        "rows": rows_out,
        "truncated": (end_row - start_row + 1) > max_rows,
    }
