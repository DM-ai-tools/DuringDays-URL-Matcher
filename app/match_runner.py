"""Run product URL matching over an Excel range with progress callbacks."""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from app.excel_io import (
    URL_HEADER_ALIASES,
    ensure_bigw_columns,
    ensure_title_column,
    row_bounds,
)
from app.sitemap_custom import load_site_index
from matcher import match_product
from product_query import ProductQuery, build_brand_set
from url_name import resolve_product_name, title_from_product_url

ProgressCb = Callable[[dict], None]


def _header_map(ws) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        mapping[str(cell.value).strip().lower()] = cell.column - 1
    return mapping


def _col(headers: dict[str, int], *names: str) -> int | None:
    for n in names:
        if n.lower() in headers:
            return headers[n.lower()]
    return None


def _write(ws, excel_row: int, col0: int | None, value) -> None:
    if col0 is None:
        return
    ws.cell(row=excel_row, column=col0 + 1, value=value)


def _accuracy(status: str, score: float, note: str) -> str:
    if status == "MATCH":
        return f"MATCH (score={score:.1f})"
    if status == "PARTIAL":
        short = note.split("best-guess=")[0].strip().rstrip(";").strip()
        if len(short) > 80:
            short = short[:77] + "..."
        return f"PARTIAL (score={score:.1f}; {short})" if short else f"PARTIAL (score={score:.1f})"
    return f"NOT FOUND (score={score:.1f})"


def _cell_str(ws, excel_row: int, col0: int | None) -> str:
    if col0 is None:
        return ""
    val = ws.cell(row=excel_row, column=col0 + 1).value
    return "" if val is None else str(val).strip()


def run_match_job(
    input_path: Path,
    output_path: Path,
    sheet: str,
    start_row: int,
    end_row: int,
    site_id: str = "bigw",
    progress: ProgressCb | None = None,
) -> dict:
    def emit(**kwargs):
        if progress:
            progress(kwargs)

    # Bounds check against actual Excel data rows (Title and/or product URL)
    bounds = row_bounds(input_path, sheet)
    if bounds["data_rows"] == 0:
        raise ValueError(
            "No data rows found. Spreadsheet needs a Title column and/or a product URL column "
            "(e.g. URL with https://www.duringdays.com.au/products/...)."
        )

    start_row = max(start_row, bounds["min_row"])
    end_row = min(end_row, bounds["max_row"])
    if start_row > end_row:
        raise ValueError(
            f"Invalid range after clamping to sheet bounds "
            f"({bounds['min_row']}–{bounds['max_row']})"
        )

    emit(
        message=f"Ensuring columns on sheet '{sheet}'",
        progress=2,
        status="running",
    )
    # Copy upload → output on first run; reuse working copy in-place on later runs.
    if input_path.resolve() != output_path.resolve():
        shutil.copy2(input_path, output_path)

    title_added = ensure_title_column(output_path, sheet)
    added = ensure_bigw_columns(output_path, sheet)
    if title_added:
        added = ["Title", *added]
    if added:
        emit(message=f"Added missing columns: {', '.join(added)}", progress=4)

    emit(message=f"Loading sitemap index '{site_id}'…", progress=8)
    records, inverted = load_site_index(site_id)
    emit(
        message=f"Index ready — {len(records):,} product URLs",
        progress=15,
    )

    wb = load_workbook(output_path)
    ws = wb[sheet]
    headers = _header_map(ws)
    COL_TITLE = _col(headers, "title")
    COL_ID = _col(headers, "id", "sku")
    COL_PRODUCT_URL = _col(headers, *URL_HEADER_ALIASES)
    COL_NAME = _col(headers, "bigw name", "bigw title")
    COL_URL = _col(headers, "bigw url")
    COL_MATCH = _col(headers, "bigw match")

    if COL_TITLE is None and COL_PRODUCT_URL is None:
        raise ValueError(
            "Spreadsheet must have a 'Title' column or a product 'URL' column "
            "(During Days product link)."
        )

    # Rebuild header map after ensure (in case columns moved)
    headers = _header_map(ws)
    COL_TITLE = _col(headers, "title")
    COL_PRODUCT_URL = _col(headers, *URL_HEADER_ALIASES)
    COL_NAME = _col(headers, "bigw name", "bigw title")
    COL_URL = _col(headers, "bigw url")
    COL_MATCH = _col(headers, "bigw match")

    # Resolve titles from product URLs where Title is blank
    titles_from_url = 0
    all_titles: list[str] = []
    for excel_row in range(2, ws.max_row + 1):
        title = _cell_str(ws, excel_row, COL_TITLE)
        product_url = _cell_str(ws, excel_row, COL_PRODUCT_URL)
        if not title and product_url.lower().startswith("http"):
            title = title_from_product_url(product_url, fetch_live=False)
            if title and COL_TITLE is not None:
                _write(ws, excel_row, COL_TITLE, title)
                titles_from_url += 1
        all_titles.append(title)

    if titles_from_url:
        emit(
            message=f"Derived {titles_from_url:,} Title values from product URLs",
            progress=17,
        )
        wb.save(output_path)

    brand_set = build_brand_set(all_titles)

    total = end_row - start_row + 1
    counters = {"MATCH": 0, "PARTIAL": 0, "NOT FOUND": 0}
    done = 0
    skipped = 0

    emit(
        message=f"Matching rows {start_row}–{end_row} ({total} rows)",
        progress=18,
        total_rows=total,
        current_row=start_row,
    )

    for excel_row in range(start_row, end_row + 1):
        title = _cell_str(ws, excel_row, COL_TITLE)
        product_url = _cell_str(ws, excel_row, COL_PRODUCT_URL)

        # If Title still empty in range, try URL again (row may have been outside pre-scan)
        if not title and product_url.lower().startswith("http"):
            title = title_from_product_url(product_url, fetch_live=False)
            if title:
                _write(ws, excel_row, COL_TITLE, title)
                titles_from_url += 1

        if not title:
            skipped += 1
            done += 1
            continue

        pq = ProductQuery.from_title(title, brand_set=brand_set)
        # For custom sites that aren't bigw/kogan, still use match_product —
        # brand filter uses inverted index the same way.
        result = match_product(
            pq, records, inverted, site_id if site_id in ("bigw", "kogan") else "bigw"
        )

        status = result.status
        out_url = result.url
        out_name = "-"
        if status == "MATCH" and out_url and out_url != "-":
            out_name = resolve_product_name(out_url, fetch_live=False)

        accuracy = _accuracy(status, result.score, result.note)
        _write(ws, excel_row, COL_NAME, out_name)
        _write(ws, excel_row, COL_URL, out_url)
        _write(ws, excel_row, COL_MATCH, accuracy)
        counters[status] = counters.get(status, 0) + 1
        done += 1

        pct = 18 + (done / max(total, 1)) * 80
        emit(
            message=f"Row {excel_row}: {status} — {title[:60]}",
            progress=round(pct, 1),
            current_row=excel_row,
            total_rows=total,
            counters=dict(counters),
            row_result={
                "excel_row": excel_row,
                "title": title[:80],
                "status": status,
                "score": round(result.score, 1),
                "url": out_url,
                "name": out_name,
                "match": accuracy,
            },
        )

        if done % 50 == 0:
            wb.save(output_path)

    wb.save(output_path)
    # Do not emit status=done here — the API sets done only after result_path is stored.
    emit(
        message="Saving workbook…",
        progress=99,
        status="running",
        counters=dict(counters),
    )
    return {
        "counters": counters,
        "start_row": start_row,
        "end_row": end_row,
        "output_path": str(output_path),
        "columns_added": added,
        "titles_from_url": titles_from_url,
        "skipped": skipped,
    }
