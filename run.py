"""CLI orchestrator: sitemaps → match → fill BigW/Kogan name+URL → write xlsx."""

from __future__ import annotations

import argparse
import csv
import shutil
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

import sitemaps
from config import (
    INPUT_XLSX as DEFAULT_INPUT_XLSX,
    OUTPUT_DIR,
    OUTPUT_XLSX,
    RUN_LOG,
    SAVE_EVERY,
    SHEET_NAME,
    SITES,
)
from matcher import match_product
from product_query import ProductQuery, build_brand_set
from url_name import resolve_product_name
from verify import verify_match


def _cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    val = row[idx].value
    return "" if val is None else str(val).strip()


def _write_cell(row, idx, value):
    if idx is not None and idx < len(row):
        row[idx].value = value


def _write_ws(ws, excel_row: int, col_0based: int | None, value) -> None:
    """Write by sheet coordinates (safe for newly added columns)."""
    if col_0based is None:
        return
    ws.cell(row=excel_row, column=col_0based + 1, value=value)


def _header_map(ws) -> dict[str, int]:
    """Map normalised header name -> 0-based column index."""
    mapping: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is None:
            continue
        key = str(cell.value).strip().lower()
        mapping[key] = cell.column - 1  # openpyxl is 1-based
    return mapping


def _col(headers: dict[str, int], *names: str) -> int | None:
    for n in names:
        if n.lower() in headers:
            return headers[n.lower()]
    return None


def _ensure_column(ws, headers: dict[str, int], name: str, after_col_0based: int | None = None) -> int:
    """Return 0-based index of column `name`, creating it if missing."""
    existing = _col(headers, name)
    if existing is not None:
        return existing

    if after_col_0based is not None:
        # Insert conceptually after a known column by writing at max(after+2, max+1)
        new_col_1based = max(after_col_0based + 2, ws.max_column + 1)
    else:
        new_col_1based = ws.max_column + 1

    ws.cell(row=1, column=new_col_1based, value=name)
    headers[name.lower()] = new_col_1based - 1
    print(f"  Added column '{name}' (col {new_col_1based})")
    return new_col_1based - 1


def _match_accuracy_label(status: str, score: float, note: str) -> str:
    """Human-readable match accuracy for the sheet."""
    if status == "MATCH":
        return f"MATCH (score={score:.1f})"
    if status == "PARTIAL":
        # Keep note short for the cell
        short = note.split("best-guess=")[0].strip().rstrip(";").strip()
        if len(short) > 80:
            short = short[:77] + "..."
        return f"PARTIAL (score={score:.1f}; {short})" if short else f"PARTIAL (score={score:.1f})"
    return f"NOT FOUND (score={score:.1f})"


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Match DuringDays products to Kogan & BigW via sitemaps"
    )
    parser.add_argument("--limit", type=int, default=None, help="Only first N data rows")
    parser.add_argument(
        "--start-row",
        type=int,
        default=None,
        help="Excel row number to start from (1=header, data starts at 2)",
    )
    parser.add_argument(
        "--end-row",
        type=int,
        default=None,
        help="Excel row number to end at (inclusive)",
    )
    parser.add_argument(
        "--refresh-sitemaps",
        action="store_true",
        help="Force re-download of sitemaps",
    )
    parser.add_argument(
        "--no-verify",
        action="store_true",
        help="Skip live JSON-LD verification (still fills Name from URL slug)",
    )
    parser.add_argument(
        "--fetch-names",
        action="store_true",
        help="Fetch live product name from matched URL (curl_cffi); "
        "falls back to slug-derived name",
    )
    parser.add_argument(
        "--site",
        choices=list(SITES),
        default=None,
        help="Only match one retailer (default: auto from available columns)",
    )
    parser.add_argument(
        "--verify",
        action="store_true",
        default=False,
        help="Enable verification (default ON unless --no-verify)",
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=None,
        help="Path to input xlsx (default: data/Final_DD_URL.xlsx)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Path to write results (default: same as --input = in-place update)",
    )
    args = parser.parse_args(argv)

    # Verification is opt-in (--verify). Name is always filled from the matched URL slug
    # (or live page when --fetch-names / --verify succeeds).
    do_verify = bool(args.verify) and not args.no_verify
    INPUT_XLSX = Path(args.input) if args.input else DEFAULT_INPUT_XLSX
    # Default: update the input file in place (Final_DD_URL workflow)
    OUT_XLSX = Path(args.output) if args.output else INPUT_XLSX

    if not INPUT_XLSX.exists():
        print(
            f"ERROR: Input not found: {INPUT_XLSX}\n"
            f"Expected data/Final_DD_URL.xlsx (sheet '{SHEET_NAME}')."
        )
        return 1

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    # Peek headers to decide which sites to run
    peek = load_workbook(INPUT_XLSX, read_only=True)
    if SHEET_NAME not in peek.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Sheets: {peek.sheetnames}")
        return 1
    headers_peek = _header_map(peek[SHEET_NAME])
    peek.close()

    has_bigw = _col(headers_peek, "bigw url", "bigw name") is not None
    has_kogan = _col(headers_peek, "kogan url", "kogan name", "kogan title") is not None

    if args.site:
        sites = [args.site]
    else:
        sites = []
        if has_bigw:
            sites.append("bigw")
        if has_kogan:
            sites.append("kogan")
        if not sites:
            sites = ["bigw"]  # fallback
    print(f"Detected columns -> bigw={has_bigw} kogan={has_kogan}; running sites={sites}")

    indexes: dict[str, tuple] = {}
    for site in sites:
        print(f"\n=== Loading {site} ===")
        records, inverted = sitemaps.load_or_build(site, refresh=args.refresh_sitemaps)
        indexes[site] = (records, inverted)
        print(f"[{site}] ready: {len(records)} URLs indexed")

    if INPUT_XLSX.resolve() == OUT_XLSX.resolve():
        wb = load_workbook(INPUT_XLSX)
        print(f"Updating in place: {INPUT_XLSX}")
    else:
        shutil.copy2(INPUT_XLSX, OUT_XLSX)
        wb = load_workbook(OUT_XLSX)

    ws = wb[SHEET_NAME]
    headers = _header_map(ws)

    COL_ID = _col(headers, "id")
    COL_TITLE = _col(headers, "title")
    COL_BIGW_NAME = _col(headers, "bigw name", "bigw title")
    COL_BIGW_URL = _col(headers, "bigw url")
    COL_KOGAN_NAME = _col(headers, "kogan name", "kogan title")
    COL_KOGAN_URL = _col(headers, "kogan url")
    COL_KOGAN_PRICE = _col(headers, "kogan price")
    COL_BIGW_PRICE = _col(headers, "bigw price")
    COL_VALIDATION = _col(headers, "validation")

    # Ensure match-accuracy columns exist (placed after Name/URL columns)
    COL_BIGW_MATCH = None
    COL_KOGAN_MATCH = None
    if "bigw" in sites or COL_BIGW_URL is not None or COL_BIGW_NAME is not None:
        after = COL_BIGW_URL if COL_BIGW_URL is not None else COL_BIGW_NAME
        COL_BIGW_MATCH = _ensure_column(ws, headers, "BigW Match", after_col_0based=after)
    if "kogan" in sites or COL_KOGAN_URL is not None or COL_KOGAN_NAME is not None:
        after = COL_KOGAN_URL if COL_KOGAN_URL is not None else COL_KOGAN_NAME
        COL_KOGAN_MATCH = _ensure_column(ws, headers, "Kogan Match", after_col_0based=after)

    if COL_TITLE is None:
        print("ERROR: No 'Title' column found in sheet")
        return 1

    print(
        f"Column map: Title={COL_TITLE} BigW Name={COL_BIGW_NAME} BigW URL={COL_BIGW_URL} "
        f"BigW Match={COL_BIGW_MATCH} Kogan Name={COL_KOGAN_NAME} Kogan URL={COL_KOGAN_URL} "
        f"Kogan Match={COL_KOGAN_MATCH}"
    )

    all_titles: list[str] = []
    for row in ws.iter_rows(min_row=2):
        all_titles.append(_cell(row, COL_TITLE))
    brand_set = build_brand_set(all_titles)
    print(f"Brand set size: {len(brand_set)}")

    log_fields = [
        "ID",
        "Excel Row",
        "Kogan Status",
        "Kogan Score",
        "Kogan Candidates",
        "Kogan URL",
        "Kogan Name",
        "BigW Status",
        "BigW Score",
        "BigW Candidates",
        "BigW URL",
        "BigW Name",
    ]
    run_log_file = open(RUN_LOG, "w", newline="", encoding="utf-8")
    log_writer = csv.DictWriter(run_log_file, fieldnames=log_fields)
    log_writer.writeheader()

    counters = {s: Counter() for s in ("kogan", "bigw")}
    rows_done = 0
    start_row = args.start_row if args.start_row is not None else 2
    end_row = args.end_row
    if start_row < 2:
        print("ERROR: --start-row must be >= 2 (row 1 is the header)")
        return 1
    print(
        f"Processing Excel rows {start_row}..{end_row or 'end'} "
        f"(verify={do_verify}, fetch_names={args.fetch_names})"
    )

    try:
        for excel_row, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if excel_row < start_row:
                continue
            if end_row is not None and excel_row > end_row:
                break
            if args.limit is not None and rows_done >= args.limit:
                break

            pid = _cell(row, COL_ID)
            title = _cell(row, COL_TITLE)
            if not title:
                print(f"  row {excel_row}: SKIP (empty title)", flush=True)
                rows_done += 1
                continue

            print(f"\n  row {excel_row}: {title[:70]}", flush=True)
            pq = ProductQuery.from_title(title, brand_set=brand_set)
            log_row = {k: "" for k in log_fields}
            log_row["ID"] = pid
            log_row["Excel Row"] = str(excel_row)
            log_row.update(
                {
                    "Kogan Status": "-",
                    "Kogan URL": "-",
                    "Kogan Name": "-",
                    "BigW Status": "-",
                    "BigW URL": "-",
                    "BigW Name": "-",
                }
            )
            validation_parts: list[str] = []

            for site in sites:
                if site == "bigw":
                    col_name, col_url, col_price = (
                        COL_BIGW_NAME,
                        COL_BIGW_URL,
                        COL_BIGW_PRICE,
                    )
                    col_match = COL_BIGW_MATCH
                    status_key, score_key, cand_key = (
                        "BigW Status",
                        "BigW Score",
                        "BigW Candidates",
                    )
                    url_key, name_key = "BigW URL", "BigW Name"
                    label = "BigW"
                else:
                    col_name, col_url, col_price = (
                        COL_KOGAN_NAME,
                        COL_KOGAN_URL,
                        COL_KOGAN_PRICE,
                    )
                    col_match = COL_KOGAN_MATCH
                    status_key, score_key, cand_key = (
                        "Kogan Status",
                        "Kogan Score",
                        "Kogan Candidates",
                    )
                    url_key, name_key = "Kogan URL", "Kogan Name"
                    label = "Kogan"

                if col_url is None and col_name is None and col_match is None:
                    print(f"    [{site}] no output columns — skip", flush=True)
                    continue

                records, inverted = indexes[site]
                result = match_product(pq, records, inverted, site)
                status = result.status
                note = result.note
                out_url = result.url  # "-" unless MATCH (or WRITE_PARTIAL_URL)
                out_name = "-"
                out_price = "-"

                if status == "MATCH" and out_url and out_url != "-":
                    # Always fill name from mapped URL (slug); optionally live-fetch
                    if do_verify:
                        print(f"    [{site}] MATCH -> verifying...", flush=True)
                        vr = verify_match(out_url, pq)
                        if vr.status == "VERIFIED":
                            note = (
                                "brand + variant attrs confirmed in URL slug; "
                                "verified via JSON-LD"
                            )
                            out_name = vr.title if vr.title != "-" else resolve_product_name(
                                out_url, fetch_live=args.fetch_names
                            )
                            out_price = vr.price
                        elif vr.status == "PARTIAL":
                            status = "PARTIAL"
                            note = vr.note
                            out_url = "-"
                            out_name = "-"
                            out_price = "-"
                        else:
                            note = f"{result.note}; {vr.note}"
                            out_name = resolve_product_name(
                                result.best_url or out_url,
                                fetch_live=args.fetch_names,
                            )
                    else:
                        out_name = resolve_product_name(
                            out_url, fetch_live=args.fetch_names
                        )

                accuracy = _match_accuracy_label(status, result.score, note)
                _write_ws(ws, excel_row, col_name, out_name)
                _write_ws(ws, excel_row, col_url, out_url)
                _write_ws(ws, excel_row, col_price, out_price)
                _write_ws(ws, excel_row, col_match, accuracy)

                if status == "MATCH":
                    validation_parts.append(f"{label}=MATCH: {note}")
                elif status == "PARTIAL":
                    validation_parts.append(f"{label}=PARTIAL: {note}")
                else:
                    validation_parts.append(f"{label}=NOT FOUND: {note}")

                counters[site][status] += 1
                log_row[status_key] = status
                log_row[score_key] = f"{result.score:.1f}"
                log_row[cand_key] = str(result.candidates_considered)
                log_row[url_key] = result.best_url if result.best_url else "-"
                log_row[name_key] = out_name

                url_show = out_url if out_url != "-" else (result.best_url or "-")
                print(
                    f"    [{site}] {accuracy}",
                    flush=True,
                )
                print(f"           URL : {url_show[:100]}", flush=True)
                print(f"           Name: {out_name[:80]}", flush=True)

            if COL_VALIDATION is not None and validation_parts:
                _write_ws(ws, excel_row, COL_VALIDATION, " | ".join(validation_parts))

            log_writer.writerow(log_row)
            run_log_file.flush()

            rows_done += 1
            if rows_done % SAVE_EVERY == 0:
                wb.save(OUT_XLSX)
                print(f"  saved checkpoint -> {OUT_XLSX}", flush=True)

        wb.save(OUT_XLSX)
    finally:
        run_log_file.close()

    print("\n========== SUMMARY ==========")
    print(f"Rows processed: {rows_done}")
    for site in sites:
        c = counters[site]
        print(
            f"  {site}: MATCH={c['MATCH']}  PARTIAL={c['PARTIAL']}  "
            f"NOT FOUND={c['NOT FOUND']}"
        )
    print(f"Output xlsx : {OUT_XLSX}")
    print(f"Run log     : {RUN_LOG}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
